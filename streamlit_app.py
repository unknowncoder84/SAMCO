"""
SAMCO Bhavcopy Downloader - Streamlit Version
Modern UI with dark theme matching the Next.js design
"""
import streamlit as st
from datetime import datetime, timedelta
import asyncio
import sys
from pathlib import Path

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent / "backend"))

from backend.app.playwright_scraper import PlaywrightSamcoScraper
from backend.app.holiday_checker import HolidayChecker
from backend.app.exceptions import HolidayException, FileNotUploadedException
import pandas as pd
from io import BytesIO
from backend.app.data_processor import DataProcessor
import json

# Page config - must be first Streamlit command
st.set_page_config(
    page_title="SAMCO Bhavcopy Downloader",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for dark theme matching Next.js design
st.markdown("""
<style>
    /* Dark theme */
    .stApp {
        background-color: #0a0a0a;
    }
    
    /* Main container */
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    
    /* Headers */
    h1 {
        color: #ffffff;
        font-weight: 700;
        margin-bottom: 0.5rem;
    }
    
    h2, h3 {
        color: #e5e5e5;
        font-weight: 600;
    }
    
    /* Cards */
    .stCard {
        background-color: #18181b;
        border: 1px solid #27272a;
        border-radius: 0.5rem;
        padding: 1.5rem;
    }
    
    /* Buttons */
    .stButton > button {
        background-color: #10b981;
        color: white;
        font-weight: 600;
        border: none;
        border-radius: 0.5rem;
        padding: 0.75rem 1.5rem;
        width: 100%;
        transition: all 0.2s;
    }
    
    .stButton > button:hover {
        background-color: #059669;
        box-shadow: 0 0 20px rgba(16, 185, 129, 0.5);
    }
    
    /* Download button */
    .stDownloadButton > button {
        background-color: #10b981;
        color: white;
        font-weight: 600;
        border: none;
        border-radius: 0.5rem;
        padding: 0.75rem 1.5rem;
        width: 100%;
    }
    
    /* Input fields */
    .stDateInput > div > div > input,
    .stSelectbox > div > div > select,
    .stMultiSelect > div > div {
        background-color: #18181b;
        border: 1px solid #27272a;
        color: #ffffff;
        border-radius: 0.5rem;
    }
    
    /* Checkboxes */
    .stCheckbox {
        color: #e5e5e5;
    }
    
    /* Success/Error messages */
    .stSuccess {
        background-color: #064e3b;
        border: 1px solid #10b981;
        color: #10b981;
    }
    
    .stError {
        background-color: #450a0a;
        border: 1px solid #ef4444;
        color: #ef4444;
    }
    
    .stWarning {
        background-color: #451a03;
        border: 1px solid #f59e0b;
        color: #f59e0b;
    }
    
    .stInfo {
        background-color: #172554;
        border: 1px solid #3b82f6;
        color: #3b82f6;
    }
    
    /* Sidebar */
    .css-1d391kg {
        background-color: #18181b;
    }
    
    /* Text */
    p, label, .stMarkdown {
        color: #d4d4d8;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background-color: #18181b;
        border: 1px solid #27272a;
        color: #ffffff;
    }
    
    /* File uploader */
    .stFileUploader {
        background-color: #18181b;
        border: 1px solid #27272a;
        border-radius: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'logs' not in st.session_state:
    st.session_state.logs = []

def add_log(level, message):
    """Add log message to session state"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    st.session_state.logs.append({
        'time': timestamp,
        'level': level,
        'message': message
    })

def download_csv(segment, date):
    """Download CSV file from Samco"""
    try:
        add_log('info', f'Starting download for {segment} on {date.strftime("%Y-%m-%d")}...')
        
        # Check if holiday
        if HolidayChecker.is_holiday(date):
            error_msg = HolidayChecker.get_holiday_message(date)
            add_log('warning', error_msg)
            return None, error_msg
        
        # Download using Playwright
        async def fetch():
            async with PlaywrightSamcoScraper(headless=True) as scraper:
                return await scraper.fetch_segment_data(segment, date)
        
        # Run async function
        csv_data = asyncio.run(fetch())
        
        if csv_data:
            add_log('success', f'Downloaded {len(csv_data):,} bytes')
            return csv_data, None
        else:
            error_msg = "No data returned from Samco"
            add_log('error', error_msg)
            return None, error_msg
            
    except FileNotUploadedException as e:
        error_msg = str(e)
        add_log('error', error_msg)
        return None, error_msg
    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}"
        add_log('error', error_msg)
        return None, error_msg

# Header
st.markdown("# 📊 SAMCO Bhavcopy Downloader")
st.markdown("Download market data files from Samco for NSE, BSE, and MCX segments")

# Sidebar
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    
    # Date selection
    st.markdown("### 📅 Select Date")
    selected_date = st.date_input(
        "Trading Date",
        value=datetime.now(),
        max_value=datetime.now(),
        help="Select the trading date to download data"
    )
    
    # Segment selection
    st.markdown("### 📈 Select Segment")
    segments = st.multiselect(
        "Market Segments",
        options=["NSE_CASH", "NSE_FO", "BSE", "MCX"],
        default=["NSE_FO"],
        help="Select one or more market segments"
    )
    
    st.markdown("---")
    
    # Info
    st.markdown("### ℹ️ About")
    st.markdown("""
    This tool downloads bhavcopy (market data) files from Samco.
    
    **Features:**
    - Direct CSV download
    - CSV to Excel conversion
    - Column filtering
    - PE/CE filtering
    """)

# Main content
tab1, tab2, tab3 = st.tabs(["📥 Download CSV", "📊 CSV to Excel", "📜 Logs"])

with tab1:
    st.markdown("## Download CSV Files")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### Selected Configuration")
        st.info(f"""
        **Date:** {selected_date.strftime('%d %B %Y')} ({selected_date.strftime('%A')})  
        **Segments:** {', '.join(segments) if segments else 'None selected'}
        """)
        
        if not segments:
            st.warning("⚠️ Please select at least one segment from the sidebar")
        
        # Download button
        if st.button("🔥 Download CSV File", disabled=not segments, use_container_width=True):
            if len(segments) == 1:
                # Single segment download
                with st.spinner(f'Downloading {segments[0]}...'):
                    csv_data, error = download_csv(segments[0], selected_date)
                    
                    if csv_data:
                        # Create filename
                        date_str = selected_date.strftime("%Y%m%d")
                        segment_map = {
                            "NSE_CASH": "NSE_CASH",
                            "NSE_FO": "NSE_FO",
                            "BSE": "BSE",
                            "MCX": "MCX"
                        }
                        filename = f"bhavcopy_{date_str}_{segment_map[segments[0]]}.csv"
                        
                        st.success(f"✅ Downloaded successfully! ({len(csv_data):,} bytes)")
                        
                        # Download button
                        st.download_button(
                            label="📥 Download File",
                            data=csv_data,
                            file_name=filename,
                            mime="text/csv",
                            use_container_width=True
                        )
                    else:
                        st.error(f"❌ {error}")
            else:
                # Multiple segments - download all
                st.info("Downloading multiple segments...")
                downloaded_files = {}
                
                for segment in segments:
                    with st.spinner(f'Downloading {segment}...'):
                        csv_data, error = download_csv(segment, selected_date)
                        
                        if csv_data:
                            date_str = selected_date.strftime("%Y%m%d")
                            filename = f"bhavcopy_{date_str}_{segment}.csv"
                            downloaded_files[filename] = csv_data
                            st.success(f"✅ {segment}: {len(csv_data):,} bytes")
                        else:
                            st.error(f"❌ {segment}: {error}")
                
                if downloaded_files:
                    # Create ZIP file
                    import zipfile
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for filename, data in downloaded_files.items():
                            zip_file.writestr(filename, data)
                    
                    zip_buffer.seek(0)
                    
                    st.download_button(
                        label=f"📥 Download All ({len(downloaded_files)} files)",
                        data=zip_buffer,
                        file_name=f"bhavcopy_{selected_date.strftime('%Y%m%d')}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
    
    with col2:
        st.markdown("### Quick Info")
        
        # Check if holiday
        if HolidayChecker.is_holiday(selected_date):
            st.error("🚫 Holiday/Weekend")
            st.caption(HolidayChecker.get_holiday_message(selected_date))
        else:
            st.success("✅ Trading Day")
        
        # Show day of week
        day_name = selected_date.strftime("%A")
        st.metric("Day", day_name)
        
        # Show segments count
        st.metric("Segments", len(segments))

with tab2:
    st.markdown("## CSV to Excel Converter")
    
    st.markdown("Upload a CSV file and convert it to Excel with column filtering and PE/CE options.")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        # File upload
        uploaded_file = st.file_uploader("Upload CSV File", type=['csv'])
        
        if uploaded_file:
            try:
                # Read CSV
                df = pd.read_csv(uploaded_file)
                
                st.success(f"✅ Loaded {len(df)} rows, {len(df.columns)} columns")
                
                # Column selection
                st.markdown("### Select Columns to Include")
                selected_columns = st.multiselect(
                    "Columns",
                    options=list(df.columns),
                    default=list(df.columns)[:5] if len(df.columns) > 5 else list(df.columns),
                    help="Select which columns to include in the Excel file"
                )
                
                # PE/CE filtering
                st.markdown("### Option Type Filtering")
                col_pe, col_ce = st.columns(2)
                with col_pe:
                    include_pe = st.checkbox("Include PE (Put)", value=True)
                with col_ce:
                    include_ce = st.checkbox("Include CE (Call)", value=True)
                
                # Convert button
                if st.button("🔄 Convert to Excel", use_container_width=True):
                    if not selected_columns:
                        st.warning("⚠️ Please select at least one column")
                    else:
                        with st.spinner("Converting to Excel..."):
                            # Filter columns
                            df_filtered = df[selected_columns].copy()
                            
                            # Filter PE/CE
                            if not include_pe:
                                mask = df_filtered.astype(str).apply(
                                    lambda x: x.str.upper().str.contains(r'\bPE\b', na=False, regex=True)
                                ).any(axis=1)
                                df_filtered = df_filtered[~mask]
                            
                            if not include_ce:
                                mask = df_filtered.astype(str).apply(
                                    lambda x: x.str.upper().str.contains(r'\bCE\b', na=False, regex=True)
                                ).any(axis=1)
                                df_filtered = df_filtered[~mask]
                            
                            # Convert to Excel
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Filtered Data"
                            
                            # Add data
                            for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
                                for c_idx, value in enumerate(row, 1):
                                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                    
                                    if r_idx == 1:
                                        # Header
                                        cell.font = Font(bold=True, color="FFFFFF")
                                        cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                    else:
                                        # Data
                                        if isinstance(value, (int, float)):
                                            if isinstance(value, float):
                                                cell.number_format = '#,##0.00'
                                            else:
                                                cell.number_format = '#,##0'
                                        cell.alignment = Alignment(horizontal="left", vertical="center")
                                    
                                    # Borders
                                    thin_border = Border(
                                        left=Side(style='thin', color='D1D5DB'),
                                        right=Side(style='thin', color='D1D5DB'),
                                        top=Side(style='thin', color='D1D5DB'),
                                        bottom=Side(style='thin', color='D1D5DB')
                                    )
                                    cell.border = thin_border
                            
                            # Auto-adjust column widths
                            for column in ws.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width
                            
                            # Freeze header
                            ws.freeze_panes = "A2"
                            
                            # Save to BytesIO
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            st.success(f"✅ Converted {len(df_filtered)} rows to Excel")
                            
                            # Download button
                            filename = uploaded_file.name.replace('.csv', '_filtered.xlsx')
                            st.download_button(
                                label="📥 Download Excel File",
                                data=excel_buffer,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    with col2:
        if uploaded_file:
            st.markdown("### File Info")
            st.metric("Rows", len(df))
            st.metric("Columns", len(df.columns))
            
            if selected_columns:
                st.metric("Selected Columns", len(selected_columns))

with tab3:
    st.markdown("## System Logs")
    
    if st.button("🗑️ Clear Logs"):
        st.session_state.logs = []
        st.rerun()
    
    if st.session_state.logs:
        for log in reversed(st.session_state.logs[-50:]):  # Show last 50 logs
            level = log['level']
            if level == 'success':
                st.success(f"[{log['time']}] ✅ {log['message']}")
            elif level == 'error':
                st.error(f"[{log['time']}] ❌ {log['message']}")
            elif level == 'warning':
                st.warning(f"[{log['time']}] ⚠️ {log['message']}")
            else:
                st.info(f"[{log['time']}] ℹ️ {log['message']}")
    else:
        st.info("No logs yet. Start downloading files to see activity logs.")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #71717a; font-size: 0.875rem;'>
    <p>SAMCO Bhavcopy Downloader | Built with Streamlit | Data from Samco</p>
</div>
""", unsafe_allow_html=True)
