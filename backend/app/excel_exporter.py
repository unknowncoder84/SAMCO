"""
Excel export module with multi-sheet support.

This module generates Excel files with proper formatting,
filter application, and column management.
"""
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """Multi-sheet Excel export with formatting"""

    @classmethod
    def export_to_excel(
        cls,
        data: Dict[str, pd.DataFrame],
        filters: Optional[Dict] = None,
        hidden_columns: Optional[Dict] = None,
        filename: Optional[str] = None,
    ) -> str:
        """
        Export data to Excel with multiple sheets.
        Format matches CSV exactly - simple and clean.

        Args:
            data: {segment: DataFrame}
            filters: {segment: [ColumnFilter]}
            hidden_columns: {segment: [column_names]}
            filename: Output filename (default: bhavcopy_DDMMMYYYY.xlsx)

        Returns:
            Path to generated file
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for segment, df in data.items():
            if df.empty:
                continue

            # Apply filters
            if filters and segment in filters:
                df = cls._apply_filters(df, filters[segment])

            # Remove hidden columns
            if hidden_columns and segment in hidden_columns:
                df = df.drop(columns=hidden_columns[segment], errors="ignore")

            # Create sheet
            ws = wb.create_sheet(title=segment)

            # Write data exactly as it appears in CSV
            # Header row
            for c_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=c_idx, value=str(col_name))

            # Data rows - write exactly as they are
            for r_idx, row in enumerate(df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    
                    # Write value as-is (like CSV)
                    if pd.isna(value):
                        cell.value = ""
                    else:
                        cell.value = value
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Generate filename
        if not filename:
            date_str = datetime.now().strftime("%d%b%Y")
            filename = f"bhavcopy_{date_str}.xlsx"

        # Save file
        wb.save(filename)
        return filename

    @classmethod
    def _apply_filters(cls, df: pd.DataFrame, filters: List[Dict]) -> pd.DataFrame:
        """
        Apply column filters to DataFrame.

        Args:
            df: Input DataFrame
            filters: List of filter definitions

        Returns:
            Filtered DataFrame
        """
        for filter_def in filters:
            field = filter_def["field"]
            values = filter_def["values"]
            if field in df.columns:
                df = df[df[field].isin(values)]
        return df
