"""
FastAPI main application entry point with complete API routes
"""
import logging
import json
from datetime import datetime
from typing import List, Optional, Dict
from fastapi import FastAPI, HTTPException, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
import pandas as pd
from io import BytesIO

from .scraper import BhavcopyScraper
from .simple_downloader import SimplePlaywrightDownloader
from .exceptions import HolidayException, FileNotUploadedException, BrowserAutomationException
from .holiday_checker import HolidayChecker
from .data_processor import DataProcessor
from .excel_exporter import ExcelExporter

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

app = FastAPI(title="Bhavcopy Pro API", version="1.0.0")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",  # Frontend on alternate port
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================================
# Pydantic Models
# ============================================================================


class ScrapeRequest(BaseModel):
    """Request model for scraping bhavcopy data"""

    segments: List[str] = Field(..., description="List of segments to scrape")
    symbols: Optional[List[str]] = Field(None, description="Optional symbol filter")
    date: Optional[str] = Field(None, description="Optional date in YYYY-MM-DD format")


class ScrapeResponse(BaseModel):
    """Response model for scraping operation"""

    success: bool
    data: Dict[str, List[Dict]]
    dates_used: Dict[str, str]
    message: str
    logs: List[str]


class ExportRequest(BaseModel):
    """Request model for Excel export"""

    data: Dict[str, List[Dict]]
    filters: Optional[Dict] = None
    hidden_columns: Optional[Dict] = None


# ============================================================================
# API Routes
# ============================================================================


@app.get("/")
async def root():
    """Root endpoint - API info"""
    return {
        "name": "Bhavcopy Pro API",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "health": "/api/health",
            "download": "/api/download-csv",
            "scrape": "/api/scrape",
            "export": "/api/export"
        }
    }


@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    logger.info("[Health] Health check requested")
    return {"status": "healthy"}


@app.post("/api/scrape", response_model=ScrapeResponse)
async def scrape_bhavcopy(request: ScrapeRequest):
    """
    Main scraping endpoint with Playwright automation.
    
    Fetches bhavcopy data for selected segments using browser automation.
    """
    logs = []
    data = {}
    dates_used = {}

    try:
        logger.info("[Scraper] Initializing...")
        logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] [Scraper] Initializing...")

        # Validate segments
        valid_segments = ["NSE_CASH", "NSE_FO", "BSE", "MCX"]
        for segment in request.segments:
            if segment not in valid_segments:
                raise HTTPException(
                    status_code=400, detail=f"Invalid segment: {segment}"
                )

        # Parse user-provided date
        if request.date:
            try:
                trading_date = datetime.strptime(request.date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            # Default to current date if not provided
            trading_date = datetime.now()

        # Check if date is a holiday BEFORE attempting to scrape
        if HolidayChecker.is_holiday(trading_date):
            error_msg = HolidayChecker.get_holiday_message(trading_date)
            logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] [Warning] {error_msg}")
            return ScrapeResponse(
                success=False,
                data={},
                dates_used={},
                message=error_msg,
                logs=logs
            )

        # Use HTTP-based scraper (works reliably on Windows)
        async with BhavcopyScraper() as scraper:
            for segment in request.segments:
                try:
                    date_str = trading_date.strftime("%d-%b-%Y")
                    dates_used[segment] = date_str

                    logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] [Network] Fetching {segment} data for {date_str}...")

                    # Fetch data using direct HTTP download
                    csv_data = await scraper.fetch_segment_data(segment, trading_date)

                    if csv_data is None:
                        # No data available (empty file or file not found)
                        if trading_date.date() == datetime.now().date():
                            logs.append(
                                f"[{datetime.now().strftime('%H:%M:%S')}] [Warning] "
                                f"Data for today ({date_str}) is not available yet. "
                                f"Samco publishes data after 6:00 PM IST."
                            )
                        else:
                            day_of_week = trading_date.strftime("%A")
                            if day_of_week in ["Saturday", "Sunday"]:
                                logs.append(
                                    f"[{datetime.now().strftime('%H:%M:%S')}] [Warning] "
                                    f"No data for {day_of_week}, {date_str}. Market is closed on weekends."
                                )
                            else:
                                logs.append(
                                    f"[{datetime.now().strftime('%H:%M:%S')}] [Warning] "
                                    f"No data found for {date_str}. This may be a holiday."
                                )
                        continue
                    
                    if csv_data:
                        # Process data with year filtering
                        df = DataProcessor.process_csv(csv_data, request.symbols, filter_year=trading_date.year)
                        
                        # Check if DataFrame is empty
                        if len(df) == 0:
                            # Check if it's today
                            if trading_date.date() == datetime.now().date():
                                logs.append(
                                    f"[{datetime.now().strftime('%H:%M:%S')}] [Warning] "
                                    f"Data for today ({date_str}) is not available yet. "
                                    f"Samco publishes data after 6:00 PM IST."
                                )
                            else:
                                logs.append(
                                    f"[{datetime.now().strftime('%H:%M:%S')}] [Warning] "
                                    f"No data found for {date_str}. This may be a holiday."
                                )
                            continue
                        
                        if segment == "NSE_FO":
                            df = DataProcessor.sort_fo_data(df)
                        df = df.fillna(value=pd.NA)
                        data[segment] = df.to_dict(orient="records")
                        logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] [Success] Found {len(df)} records for {date_str}")

                except Exception as e:
                    error_msg = f"Error fetching {segment}: {str(e)}"
                    logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] [Warning] {error_msg}")
                    # Continue with other segments
                
                except BrowserAutomationException as e:
                    error_msg = f"Browser automation failed for {segment}: {str(e)}"
                    logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] [Error] {error_msg}")
                    # Continue with other segments

        return ScrapeResponse(
            success=True,
            data=data,
            dates_used=dates_used,
            message="Scraping completed successfully",
            logs=logs,
        )

    except Exception as e:
        error_msg = f"Scraping failed: {str(e)}"
        logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] [Error] {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)


@app.post("/api/download-csv")
async def download_csv_direct(request: ScrapeRequest):
    """
    Direct CSV download endpoint - downloads raw CSV files without processing.
    
    Returns the CSV file directly for download.
    """
    print(f"[ENDPOINT] download_csv_direct called!")
    print(f"[ENDPOINT] Request: date={request.date}, segments={request.segments}")
    
    try:
        logger.info(f"[Download] Request received: date={request.date}, segments={request.segments}")
        
        # Validate segments
        valid_segments = ["NSE_CASH", "NSE_FO", "BSE", "MCX"]
        for segment in request.segments:
            if segment not in valid_segments:
                raise HTTPException(
                    status_code=400, detail=f"Invalid segment: {segment}"
                )

        # Parse user-provided date
        if request.date:
            try:
                trading_date = datetime.strptime(request.date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            trading_date = datetime.now()

        logger.info(f"[Download] Parsed date: {trading_date.strftime('%Y-%m-%d')}")
        
        # Single segment - return CSV directly
        segment = request.segments[0]
        logger.info(f"[Download] Fetching data for {segment} using subprocess...")
        
        # Use subprocess-based downloader to avoid event loop issues
        csv_data = SimplePlaywrightDownloader.download(segment, trading_date)
        
        # Log what we received
        data_size = len(csv_data) if csv_data else 0
        logger.info(f"[Download] Received {data_size} bytes from scraper")
        
        if not csv_data or data_size == 0:
            raise FileNotUploadedException(
                f"No data available for {segment} on {trading_date.strftime('%d-%b-%Y')}"
            )
        
        # Create filename
        date_str = trading_date.strftime("%Y%m%d")
        segment_map = {
            "NSE_CASH": "NSE_CASH",
            "NSE_FO": "NSE_FO",
            "BSE": "BSE",
            "MCX": "MCX",
        }
        filename = f"bhavcopy_{date_str}_{segment_map[segment]}.csv"
        
        # Return file
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        logger.info(f"[Download] Returning {data_size} bytes as {filename}")
        
        return StreamingResponse(
            BytesIO(csv_data),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except FileNotUploadedException as e:
        error_msg = str(e) if str(e) else "File not available on Samco"
        logger.warning(f"[FileNotUploaded] {error_msg}")
        raise HTTPException(status_code=404, detail=error_msg)
    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}" if str(e) else f"{type(e).__name__} occurred"
        logger.error(f"[Error] {error_msg}")
        import traceback
        logger.error(f"[Traceback] {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=error_msg)


@app.post("/api/export")
async def export_excel(request: ExportRequest):
    """
    Excel export endpoint.

    Generates a multi-sheet Excel file with filtered data.
    """
    try:
        logger.info("[Export] Starting Excel export...")

        # Convert data back to DataFrames
        dataframes = {}
        for segment, records in request.data.items():
            if records:
                df = pd.DataFrame(records)
                # Ensure numeric columns are properly typed
                for col in df.columns:
                    try:
                        # Try to convert to numeric
                        df[col] = pd.to_numeric(df[col], errors='ignore')
                    except:
                        pass
                dataframes[segment] = df

        # Export to Excel
        filename = ExcelExporter.export_to_excel(
            dataframes, request.filters, request.hidden_columns
        )

        logger.info(f"[Export] Excel file generated: {filename}")

        return FileResponse(
            filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    except Exception as e:
        error_msg = f"Export failed: {str(e)}"
        logger.error(f"[Error] {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)


@app.post("/api/process-csv-to-excel")
async def process_csv_to_excel(
    file: UploadFile = File(...),
    columns: str = Form(...),
    include_pe: str = Form(default="true"),
    include_ce: str = Form(default="true")
):
    """
    Process uploaded CSV file and convert to Excel with selected columns.
    
    - Reads CSV file
    - Filters to selected columns only
    - Optionally excludes rows containing PE or CE based on checkboxes
    - Converts numeric columns properly
    - Formats as Excel with proper number formatting
    - Returns Excel file for download
    """
    try:
        logger.info(f"[Process] Processing CSV file: {file.filename}")
        
        # Parse selected columns
        selected_columns = json.loads(columns)
        logger.info(f"[Process] Selected columns: {len(selected_columns)}")
        
        # Parse include flags
        should_include_pe = include_pe.lower() == "true"
        should_include_ce = include_ce.lower() == "true"
        logger.info(f"[Process] Include PE: {should_include_pe}, Include CE: {should_include_ce}")
        
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(BytesIO(contents))
        
        logger.info(f"[Process] Original data: {len(df)} rows, {len(df.columns)} columns")
        
        # Filter out PE rows if not included
        if not should_include_pe:
            original_count = len(df)
            # Check all columns for "PE" (case-insensitive, as whole word or value)
            mask = df.astype(str).apply(
                lambda x: x.str.upper().str.contains(r'\bPE\b', na=False, regex=True)
            ).any(axis=1)
            df = df[~mask]
            excluded_count = original_count - len(df)
            logger.info(f"[Process] Excluded {excluded_count} rows containing PE")
        
        # Filter out CE rows if not included
        if not should_include_ce:
            original_count = len(df)
            # Check all columns for "CE" (case-insensitive, as whole word or value)
            mask = df.astype(str).apply(
                lambda x: x.str.upper().str.contains(r'\bCE\b', na=False, regex=True)
            ).any(axis=1)
            df = df[~mask]
            excluded_count = original_count - len(df)
            logger.info(f"[Process] Excluded {excluded_count} rows containing CE")
        
        # Filter to selected columns only
        available_columns = [col for col in selected_columns if col in df.columns]
        df_filtered = df[available_columns].copy()
        
        logger.info(f"[Process] Filtered data: {len(df_filtered)} rows, {len(df_filtered.columns)} columns")
        
        # Convert numeric columns properly
        for col in df_filtered.columns:
            try:
                # Try to convert to numeric, keeping NaN for non-numeric values
                df_filtered[col] = pd.to_numeric(df_filtered[col], errors='ignore')
            except:
                pass
        
        # Create Excel file with proper formatting
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Data"
        
        # Add data to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatting
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Data formatting
                    if isinstance(value, (int, float)):
                        # Number formatting
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                    
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin', color='D1D5DB'),
                    right=Side(style='thin', color='D1D5DB'),
                    top=Side(style='thin', color='D1D5DB'),
                    bottom=Side(style='thin', color='D1D5DB')
                )
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename
        original_name = file.filename.replace('.csv', '')
        excel_filename = f"{original_name}_filtered.xlsx"
        
        logger.info(f"[Process] Excel file created: {excel_filename}")
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={excel_filename}"}
        )
        
    except Exception as e:
        error_msg = f"Processing failed: {str(e)}"
        logger.error(f"[Error] {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)
